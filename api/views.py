from rest_framework import viewsets, permissions, filters
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework import permissions
from rest_framework.views import APIView
from rest_framework import generics
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db import transaction
from rest_framework.filters import SearchFilter
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import viewsets, status
from django.utils import timezone
import re, json
from urllib import request as urlrequest
from urllib.error import URLError, HTTPError
from django.contrib.auth import get_user_model
from rest_framework import viewsets, permissions, parsers

# 👉 novos imports p/ importação XLSX
from openpyxl import load_workbook
from openpyxl.utils import datetime as openpyxl_datetime
from decimal import Decimal, InvalidOperation
from datetime import datetime

from .models import (
    CustomUser,
    Entidade,
    Orgao,
    ProcessoLicitatorio,
    Lote,
    Item,
    Fornecedor,
    FornecedorProcesso,
    ItemFornecedor,
    ContratoEmpenho
)

from .serializers import (
    UserSerializer,
    EntidadeSerializer,
    OrgaoSerializer,
    ProcessoLicitatorioSerializer,
    LoteSerializer,
    ItemSerializer,
    FornecedorSerializer,
    FornecedorProcessoSerializer,
    ItemFornecedorSerializer,
    ContratoEmpenhoSerializer,
    UsuarioSerializer
)

User = get_user_model()


class UsuarioViewSet(viewsets.ModelViewSet):
    """
    CRUD de usuários do sistema.
    Acesso restrito a staff/admin.
    """
    queryset = User.objects.all().order_by("id")
    serializer_class = UsuarioSerializer
    permission_classes = [permissions.IsAdminUser]  # só staff/admin acessa
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    parser_classes = [parsers.MultiPartParser, parsers.FormParser, parsers.JSONParser]
    search_fields = ["username", "email", "first_name", "last_name"]
    ordering_fields = ["id", "username", "email", "first_name", "last_name", "last_login", "date_joined"]
    ordering = ["username"]


# ============================================================
# 1️⃣ ENTIDADE / ÓRGÃO
# ============================================================

class EntidadeViewSet(viewsets.ModelViewSet):
    queryset = Entidade.objects.all().order_by('nome')
    serializer_class = EntidadeSerializer
    permission_classes = [IsAuthenticated]


def _normalize(txt: str) -> str:
    """
    Normalização genérica para comparação de textos:
    - strip
    - troca "_" por espaço
    - remove acentos
    - upper
    - compacta espaços múltiplos
    """
    import unicodedata, re

    if not txt:
        return ""

    s = str(txt).strip()
    s = s.replace("_", " ")
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    s = s.upper()
    s = re.sub(r"\s+", " ", s)
    return s


class OrgaoViewSet(viewsets.ModelViewSet):
    serializer_class = OrgaoSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]  # pode ficar, mas não dependemos dele
    filterset_fields = ['entidade']

    # ✅ filtro garantido via código (independente do DjangoFilterBackend)
    def get_queryset(self):
        qs = Orgao.objects.select_related('entidade').order_by('nome')
        entidade_id = self.request.query_params.get('entidade')
        if entidade_id:
            qs = qs.filter(entidade_id=entidade_id)
        return qs

    # ====== sua action importar_pncp permanece igual, apenas mantida aqui ======
    @action(detail=False, methods=['post'], url_path='importar-pncp')
    def importar_pncp(self, request):
        raw_cnpj = (request.data.get('cnpj') or '').strip()
        cnpj_digits = re.sub(r'\D', '', raw_cnpj)
        if len(cnpj_digits) != 14:
            return Response({"detail": "CNPJ inválido. Informe 14 dígitos."},
                            status=status.HTTP_400_BAD_REQUEST)

        url = f"https://pncp.gov.br/api/pncp/v1/orgaos/{cnpj_digits}/unidades"

        try:
            with urlrequest.urlopen(url, timeout=20) as resp:
                if resp.status != 200:
                    return Response({"detail": f"PNCP respondeu {resp.status}"},
                                    status=status.HTTP_502_BAD_GATEWAY)
                data = json.loads(resp.read().decode('utf-8'))
        except HTTPError as e:
            return Response({"detail": f"Falha ao consultar PNCP: HTTP {e.code}"},
                            status=status.HTTP_502_BAD_GATEWAY)
        except URLError:
            return Response({"detail": "Não foi possível alcançar o PNCP."},
                            status=status.HTTP_502_BAD_GATEWAY)
        except Exception as e:
            return Response({"detail": f"Erro inesperado ao consultar PNCP: {e}"},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        if not isinstance(data, list) or not data:
            return Response({"detail": "Nenhuma unidade retornada pelo PNCP."},
                            status=status.HTTP_404_NOT_FOUND)

        ALLOW_KEYWORDS = ["SECRETARIA", "FUNDO", "CONTROLADORIA", "GABINETE"]
        EXCLUDE_KEYWORDS = ["PREFEITURA"]
        EXCLUDE_CODES = {"000000001", "000000000", "1"}

        def deve_incluir(nome_unidade: str, codigo_unidade: str) -> bool:
            n = _normalize(nome_unidade or "")
            c = (codigo_unidade or "").strip()
            if c in EXCLUDE_CODES:
                return False
            if any(word in n for word in EXCLUDE_KEYWORDS):
                return False
            return any(word in n for word in ALLOW_KEYWORDS)

        def find_entidade_by_cnpj_digits(digits: str):
            for ent in Entidade.objects.all():
                ent_digits = re.sub(r'\D', '', ent.cnpj or '')
                if ent_digits == digits:
                    return ent
            return None

        razao = (data[0].get('orgao') or {}).get('razaoSocial') or ''
        ano_atual = timezone.now().year

        with transaction.atomic():
            entidade = find_entidade_by_cnpj_digits(cnpj_digits)
            if not entidade:
                entidade = Entidade.objects.create(
                    nome=razao or f"Entidade {cnpj_digits}",
                    cnpj=cnpj_digits,
                    ano=ano_atual
                )
            else:
                if razao and entidade.nome != razao:
                    entidade.nome = razao
                    entidade.save(update_fields=['nome'])

            created, updated, ignorados = 0, 0, 0

            for u in data:
                codigo = (u.get('codigoUnidade') or '').strip()
                nome = (u.get('nomeUnidade') or '').strip()
                if not deve_incluir(nome, codigo):
                    ignorados += 1
                    continue

                orgao = None
                if codigo:
                    orgao = Orgao.objects.filter(entidade=entidade, codigo_unidade=codigo).first()
                if not orgao:
                    orgao = Orgao.objects.filter(entidade=entidade, nome__iexact=nome).first()

                if orgao:
                    changed = False
                    if codigo and orgao.codigo_unidade != codigo:
                        orgao.codigo_unidade = codigo
                        changed = True
                    if orgao.nome != nome:
                        orgao.nome = nome
                        changed = True
                    if changed:
                        orgao.save(update_fields=['nome', 'codigo_unidade'])
                        updated += 1
                else:
                    Orgao.objects.create(
                        entidade=entidade,
                        nome=nome,
                        codigo_unidade=codigo or None
                    )
                    created += 1

            orgaos_entidade = Orgao.objects.filter(entidade=entidade).order_by('nome')
            return Response({
                "entidade": {"id": entidade.id, "nome": entidade.nome, "cnpj": entidade.cnpj, "ano": entidade.ano},
                "created": created, "updated": updated, "ignored": ignorados,
                "total_orgaos_entidade": orgaos_entidade.count(),
                "orgaos": OrgaoSerializer(orgaos_entidade, many=True).data
            }, status=status.HTTP_200_OK)


# ============================================================
# 2️⃣ FORNECEDOR
# ============================================================

class FornecedorViewSet(viewsets.ModelViewSet):
    queryset = Fornecedor.objects.all().order_by('razao_social')
    serializer_class = FornecedorSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [SearchFilter, DjangoFilterBackend]
    search_fields = ['razao_social', 'cnpj']
    filterset_fields = ['cnpj']


# ============================================================
# 3️⃣ PROCESSO LICITATÓRIO
# ============================================================

class ProcessoLicitatorioViewSet(viewsets.ModelViewSet):
    queryset = (
        ProcessoLicitatorio.objects
        .select_related("entidade", "orgao")
        .all()
        .order_by("-data_abertura")
    )
    serializer_class = ProcessoLicitatorioSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter]
    search_fields = ["numero_processo", "numero_certame", "objeto"]
    filterset_fields = ["modalidade", "situacao", "entidade", "orgao"]

    # ======================================================================
    # IMPORTAÇÃO XLSX (PLANILHA PADRÃO - CADASTRO INICIAL)
    # ======================================================================
    @action(
        detail=False,
        methods=["post"],
        url_path="importar-xlsx",
        parser_classes=[parsers.MultiPartParser, parsers.FormParser],
    )
    def importar_xlsx(self, request):
        arquivo = request.FILES.get("arquivo")
        if not arquivo:
            return Response({"detail": "Envie um arquivo XLSX no campo 'arquivo'."}, status=400)
        if not arquivo.name.lower().endswith(".xlsx"):
            return Response({"detail": "O arquivo deve ser .xlsx."}, status=400)

        try:
            wb = load_workbook(arquivo, data_only=True)
        except Exception:
            return Response({"detail": "Erro ao ler o arquivo XLSX."}, status=400)

        # ------------------------ helpers ------------------------

        import unicodedata, re
        from openpyxl.utils.datetime import from_excel as excel_to_datetime

        def normalize(s):
            if s is None:
                return ""
            s = str(s).strip()
            s = unicodedata.normalize("NFD", s)
            s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
            s = s.upper()
            s = re.sub(r"\s+", " ", s)
            return s

        def to_decimal(v):
            if v in (None, ""):
                return None
            try:
                sv = str(v)
                if "," in sv and "." in sv:
                    sv = sv.replace(".", "").replace(",", ".")
                else:
                    sv = sv.replace(",", ".")
                return Decimal(sv)
            except Exception:
                return None

        def to_date(v):
            if not v:
                return None
            if isinstance(v, datetime):
                return v.date()
            if isinstance(v, (int, float)):
                try:
                    return excel_to_datetime(v, wb.epoch).date()
                except Exception:
                    return None
            if isinstance(v, str):
                txt = v.strip()
                for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
                    try:
                        return datetime.strptime(txt, fmt).date()
                    except Exception:
                        continue
            return None

        def to_datetime(v):
            if not v:
                return None
            if isinstance(v, datetime):
                return v
            if isinstance(v, (int, float)):
                try:
                    return excel_to_datetime(v, wb.epoch)
                except Exception:
                    return None
            if isinstance(v, str):
                txt = v.strip()
                for fmt in ("%d/%m/%Y %H:%M", "%d/%m/%Y", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
                    try:
                        return datetime.strptime(txt, fmt)
                    except Exception:
                        continue
            return None

        def get(ws, coord):
            v = ws[coord].value
            return "" if v is None else v

        # ------------------------ planilha principal ------------------------

        ws = None
        for name in wb.sheetnames:
            if normalize(name).startswith("CADASTRO INICIAL"):
                ws = wb[name]
                break
        if ws is None:
            ws = wb[wb.sheetnames[0]]

        # Mapeamentos de amparo legal (rótulo Excel -> value interno)
        AMPARO_EXCEL_TO_VALUE = {
            "ART. 23": "art_23",
            "ART. 24": "art_24",
            "ART. 25": "art_25",
            "ART. 4º": "art_4",
            "ART. 4O": "art_4",
            "ART. 5º": "art_5",
            "ART. 5O": "art_5",
            "ART. 28, INCISO I": "art_28_i",
            "ART. 28, INCISO II": "art_28_ii",
            "ART. 75, § 7º": "art_75_par7",
            "ART. 75, § 7O": "art_75_par7",
            "ART. 75, INCISO I": "art_75_i",
            "ART. 75, INCISO II": "art_75_ii",
            "ART. 75, INCISO III, A": "art_75_iii_a",
            "ART. 75, INCISO III, B": "art_75_iii_b",
            "ART. 75, INCISO IV, A": "art_75_iv_a",
            "ART. 75, INCISO IV, B": "art_75_iv_b",
            "ART. 75, INCISO IV, C": "art_75_iv_c",
            "ART. 75, INCISO IV, D": "art_75_iv_d",
            "ART. 75, INCISO IV, E": "art_75_iv_e",
            "ART. 75, INCISO IV, F": "art_75_iv_f",
            "ART. 75, INCISO IV, J": "art_75_iv_j",
            "ART. 75, INCISO IV, K": "art_75_iv_k",
            "ART. 75, INCISO IV, M": "art_75_iv_m",
            "ART. 75, INCISO IX": "art_75_ix",
            "ART. 75, INCISO VIII": "art_75_viii",
            "ART. 75, INCISO XV": "art_75_xv",
            "LEI 11.947/2009, ART. 14, § 1º": "lei_11947_art14_1",
            "LEI 11.947/2009, ART. 14, § 1O": "lei_11947_art14_1",
            "ART. 79, INCISO I": "art_79_i",
            "ART. 79, INCISO II": "art_79_ii",
            "ART. 79, INCISO III": "art_79_iii",
            "ART. 74, CAPUT": "art_74_caput",
            "ART. 74, I": "art_74_i",
            "ART. 74, II": "art_74_ii",
            "ART. 74, III, A": "art_74_iii_a",
            "ART. 74, III, B": "art_74_iii_b",
            "ART. 74, III, C": "art_74_iii_c",
            "ART. 74, III, D": "art_74_iii_d",
            "ART. 74, III, E": "art_74_iii_e",
            "ART. 74, III, F": "art_74_iii_f",
            "ART. 74, III, G": "art_74_iii_g",
            "ART. 74, III, H": "art_74_iii_h",
            "ART. 74, IV": "art_74_iv",
            "ART. 74, V": "art_74_v",
            "ART. 86, § 2º": "art_86_2",
            "ART. 86, § 2O": "art_86_2",
        }
        AMPARO_EXCEL_NORMALIZED = {normalize(k): v for k, v in AMPARO_EXCEL_TO_VALUE.items()}

        # Modalidade / classificação / tipo organização mapeados pelos labels
        MODALIDADE_EXCEL_NORMALIZED = {
            normalize("Pregão Eletrônico"): "Pregão Eletrônico",
            normalize("Concorrência Eletrônica"): "Concorrência Eletrônica",
            normalize("Dispensa Eletrônica"): "Dispensa Eletrônica",
            normalize("Inexigibilidade Eletrônica"): "Inexigibilidade Eletrônica",
            normalize("Adesão a Registro de Preços"): "Adesão a Registro de Preços",
            normalize("Credenciamento"): "Credenciamento",
        }
        CLASSIFICACAO_EXCEL_NORMALIZED = {
            normalize("Compras"): "Compras",
            normalize("Serviços Comuns"): "Serviços Comuns",
            normalize("Serviços de Engenharia Comuns"): "Serviços de Engenharia Comuns",
            normalize("Obras Comuns"): "Obras Comuns",
        }
        TIPO_ORG_EXCEL_NORMALIZED = {
            normalize("Lote"): "Lote",
            normalize("Item"): "Item",
        }

        # ---------- meta do processo (layout fixo da planilha padrão) ----------

        numero_processo = str(get(ws, "B7")).strip()
        data_processo_raw = get(ws, "C7")
        numero_certame = str(get(ws, "D7")).strip()
        data_certame_raw = get(ws, "E7")
        entidade_nome = str(get(ws, "G7") or "").strip()
        orgao_nome = str(get(ws, "H7") or "").strip()
        valor_global_raw = get(ws, "I7")
        objeto_raw = get(ws, "B7")

        modalidade_raw = get(ws, "A11")
        tipo_disputa_raw = get(ws, "B11")
        registro_preco_raw = get(ws, "C11")
        tipo_organizacao_raw = get(ws, "D11")
        criterio_julgamento_raw = get(ws, "E11")
        classificacao_raw = get(ws, "F11")
        fundamentacao_raw = get(ws, "G11")
        amparo_legal_raw = get(ws, "H11")
        vigencia_raw = get(ws, "I11")

        # ---------- tabela de itens (linha 15 cabeçalho, 16+ dados) ----------

        row_header = 16
        col_map = {
            "LOTE": 1,
            "N ITEM": 2,
            "DESCRICAO DO ITEM": 3,
            "ESPECIFICACAO": 4,
            "QUANTIDADE": 5,
            "UNIDADE": 6,
            "NATUREZA / DESPESA": 7,
            "VALOR REFERENCIA UNITARIO": 8,
            "CNPJ DO FORNECEDOR": 9,
        }

        def col(key):
            return col_map[key]

        itens_data = []
        for row in range(row_header + 1, ws.max_row + 1):
            desc = ws.cell(row=row, column=col("DESCRICAO DO ITEM")).value
            if not desc:
                continue
            itens_data.append(
                {
                    "descricao": desc,
                    "especificacao": ws.cell(row=row, column=col("ESPECIFICACAO")).value,
                    "quantidade": ws.cell(row=row, column=col("QUANTIDADE")).value,
                    "unidade": ws.cell(row=row, column=col("UNIDADE")).value,
                    "natureza": ws.cell(row=row, column=col("NATUREZA / DESPESA")).value,
                    "valor_referencia": ws.cell(row=row, column=col("VALOR REFERENCIA UNITARIO")).value,
                    "lote": ws.cell(row=row, column=col("LOTE")).value,
                    "cnpj": ws.cell(row=row, column=col("CNPJ DO FORNECEDOR")).value,
                }
            )

        if not itens_data:
            return Response({"detail": "Nenhum item encontrado."}, status=400)

        # ---------- fornecedores (aba opcional) ----------

        fornecedores = {}

        for name in wb.sheetnames:
            if "FORNECEDOR" not in normalize(name):
                continue
            ws_f = wb[name]
            header = None
            cols = {}

            for r in range(1, ws_f.max_row + 1):
                temp = {}
                for c in range(1, ws_f.max_column + 1):
                    v = ws_f.cell(r, c).value
                    if v:
                        temp[normalize(v)] = c
                if "CNPJ" in temp and "RAZAO SOCIAL" in temp:
                    header = r
                    cols = temp
                    break

            if not header:
                continue

            for r in range(header + 1, ws_f.max_row + 1):
                cnpj_raw = ws_f.cell(r, cols["CNPJ"]).value
                razao_raw = ws_f.cell(r, cols["RAZAO SOCIAL"]).value
                if not cnpj_raw:
                    continue
                cnpj = re.sub(r"\D", "", str(cnpj_raw))
                if len(cnpj) != 14:
                    continue
                fornecedores[cnpj] = razao_raw or cnpj
            break

        # ---------- criação no banco ----------

        with transaction.atomic():
            entidade = (
                Entidade.objects.filter(nome__iexact=entidade_nome).first()
                if entidade_nome
                else None
            )
            orgao = None
            if orgao_nome:
                qs_or = Orgao.objects.all()
                if entidade:
                    qs_or = qs_or.filter(entidade=entidade)
                orgao = qs_or.filter(nome__iexact=orgao_nome).first()

            # modalidade / classificação / tipo org normalizados
            mod_txt = None
            if modalidade_raw not in (None, ""):
                mod_txt = MODALIDADE_EXCEL_NORMALIZED.get(
                    normalize(modalidade_raw), str(modalidade_raw).strip()
                )

            class_txt = None
            if classificacao_raw not in (None, ""):
                class_txt = CLASSIFICACAO_EXCEL_NORMALIZED.get(
                    normalize(classificacao_raw), str(classificacao_raw).strip()
                )

            org_txt = None
            if tipo_organizacao_raw not in (None, ""):
                org_txt = TIPO_ORG_EXCEL_NORMALIZED.get(
                    normalize(tipo_organizacao_raw), str(tipo_organizacao_raw).strip()
                )

            situacao_txt = "Em Pesquisa"

            # modo de disputa
            modo_disputa_txt = ""
            if tipo_disputa_raw not in (None, ""):
                s = str(tipo_disputa_raw).strip().lower()
                if "aberto" in s and "fechado" in s:
                    modo_disputa_txt = "aberto_e_fechado"
                elif "aberto" in s:
                    modo_disputa_txt = "aberto"
                elif "fechado" in s:
                    modo_disputa_txt = "fechado"

            # critério julgamento
            criterio_txt = ""
            if criterio_julgamento_raw not in (None, ""):
                cj = str(criterio_julgamento_raw).strip().lower()
                if "menor" in cj:
                    criterio_txt = "menor_preco"
                elif "maior" in cj:
                    criterio_txt = "maior_desconto"

            # fundamentação (lei_14133 / lei_8666 / lei_10520)
            fundamentacao_txt = None
            if fundamentacao_raw not in (None, ""):
                f = str(fundamentacao_raw).strip()
                f_lower = f.lower()
                digits = "".join(ch for ch in f if ch.isdigit())
                if "14133" in digits or "14133" in f_lower:
                    fundamentacao_txt = "lei_14133"
                elif "8666" in digits or "8666" in f_lower:
                    fundamentacao_txt = "lei_8666"
                elif "10520" in digits or "10520" in f_lower:
                    fundamentacao_txt = "lei_10520"
                elif f in ("lei_14133", "lei_8666", "lei_10520"):
                    fundamentacao_txt = f

            # amparo legal -> value interno (art_xxx)
            amparo_legal_txt = None
            if amparo_legal_raw not in (None, ""):
                a = str(amparo_legal_raw).strip()
                if a in AMPARO_EXCEL_TO_VALUE.values():
                    amparo_legal_txt = a
                else:
                    a_norm = normalize(a)
                    amparo_legal_txt = AMPARO_EXCEL_NORMALIZED.get(a_norm, a)

            processo = ProcessoLicitatorio.objects.create(
                numero_processo=numero_processo or None,
                numero_certame=numero_certame or None,
                objeto=str(objeto_raw or "").strip(),
                modalidade=mod_txt or None,
                classificacao=class_txt or None,
                tipo_organizacao=org_txt or None,
                situacao=situacao_txt,
                data_processo=to_date(data_processo_raw),
                data_abertura=to_datetime(data_certame_raw),
                valor_referencia=to_decimal(valor_global_raw),
                vigencia_meses=int(str(vigencia_raw).split()[0]) if vigencia_raw else None,
                registro_preco=str(registro_preco_raw or "").strip().lower() in ("sim", "s"),
                entidade=entidade,
                orgao=orgao,
                fundamentacao=fundamentacao_txt,
                amparo_legal=amparo_legal_txt,
                modo_disputa=modo_disputa_txt or None,
                criterio_julgamento=criterio_txt or None,
            )

            # lotes
            lotes = {}
            for it in itens_data:
                lote_num = it["lote"]
                if lote_num and lote_num not in lotes:
                    lotes[lote_num] = Lote.objects.create(
                        processo=processo,
                        numero=lote_num,
                        descricao=f"Lote {lote_num}",
                    )

            # itens + fornecedores
            ordem = 0
            fornecedores_vinculados = set()

            for it in itens_data:
                ordem += 1
                lote_obj = lotes.get(it["lote"])
                fornecedor = None
                cnpj_raw = it.get("cnpj")
                if cnpj_raw:
                    cnpj_digits = re.sub(r"\D", "", str(cnpj_raw))
                    if len(cnpj_digits) == 14:
                        fornecedor, _ = Fornecedor.objects.get_or_create(
                            cnpj=cnpj_digits,
                            defaults={"razao_social": fornecedores.get(cnpj_digits) or cnpj_digits},
                        )
                        fornecedores_vinculados.add(fornecedor.id)

                Item.objects.create(
                    processo=processo,
                    lote=lote_obj,
                    descricao=it["descricao"],
                    quantidade=to_decimal(it["quantidade"]),
                    unidade=str(it["unidade"] or "").strip(),
                    valor_estimado=to_decimal(it["valor_referencia"]),
                    ordem=ordem,
                    fornecedor=fornecedor,
                )
                if fornecedor:
                    FornecedorProcesso.objects.get_or_create(
                        processo=processo,
                        fornecedor=fornecedor,
                    )

        return Response(
            {
                "detail": "Importação concluída.",
                "processo": self.get_serializer(processo).data,
                "lotes_criados": len(lotes),
                "itens_importados": len(itens_data),
                "fornecedores_vinculados": len(fornecedores_vinculados),
            },
            status=201,
        )

    # ======================================================================
    # ITENS DO PROCESSO
    # ======================================================================
    @action(detail=True, methods=["get"])
    def itens(self, request, *args, **kwargs):
        processo = self.get_object()
        itens = (
            Item.objects.filter(processo=processo)
            .select_related("lote", "fornecedor")
            .order_by("ordem", "id")
        )
        serializer = ItemSerializer(itens, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    # ======================================================================
    # FORNECEDORES (VÍNCULO)
    # ======================================================================
    @action(detail=True, methods=["post"], url_path="adicionar_fornecedor")
    def adicionar_fornecedor(self, request, *args, **kwargs):
        processo = self.get_object()
        fornecedor_id = request.data.get("fornecedor_id")

        if not fornecedor_id:
            return Response(
                {"error": "fornecedor_id é obrigatório."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            fornecedor = Fornecedor.objects.get(id=fornecedor_id)
        except Fornecedor.DoesNotExist:
            return Response(
                {"error": "Fornecedor não encontrado."},
                status=status.HTTP_404_NOT_FOUND,
            )

        with transaction.atomic():
            obj, created = FornecedorProcesso.objects.get_or_create(
                processo=processo, fornecedor=fornecedor
            )
        return Response(
            {
                "detail": "Fornecedor vinculado ao processo com sucesso!",
                "fornecedor": FornecedorSerializer(fornecedor).data,
                "created": created,
            },
            status=status.HTTP_201_CREATED,
        )

    @action(detail=True, methods=["get"], url_path="fornecedores")
    def fornecedores(self, request, *args, **kwargs):
        processo = self.get_object()
        fornecedores = Fornecedor.objects.filter(
            processos__processo=processo
        ).order_by("razao_social")
        serializer = FornecedorSerializer(fornecedores, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"], url_path="remover_fornecedor")
    def remover_fornecedor(self, request, *args, **kwargs):
        processo = self.get_object()
        fornecedor_id = request.data.get("fornecedor_id")

        if not fornecedor_id:
            return Response(
                {"error": "fornecedor_id é obrigatório."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        deleted, _ = FornecedorProcesso.objects.filter(
            processo=processo, fornecedor_id=fornecedor_id
        ).delete()

        if deleted:
            return Response(
                {"detail": "Fornecedor removido com sucesso."},
                status=status.HTTP_200_OK,
            )
        return Response(
            {"detail": "Nenhum vínculo encontrado para remover."},
            status=status.HTTP_404_NOT_FOUND,
        )

    # ======================================================================
    # LOTES (NESTED)
    # ======================================================================
    @action(detail=True, methods=["get", "post"], url_path="lotes")
    def lotes(self, request, *args, **kwargs):
        processo = self.get_object()

        if request.method.lower() == "get":
            qs = processo.lotes.order_by("numero")
            return Response(LoteSerializer(qs, many=True).data)

        payload = request.data

        def next_numero():
            ultimo = processo.lotes.order_by("-numero").first()
            return (ultimo.numero + 1) if ultimo else 1

        created = []
        with transaction.atomic():
            if isinstance(payload, list):
                for item in payload:
                    n = item.get("numero") or next_numero()
                    d = item.get("descricao", "")
                    obj = Lote.objects.create(
                        processo=processo, numero=n, descricao=d
                    )
                    created.append(obj)

            elif isinstance(payload, dict) and "quantidade" in payload:
                try:
                    qtd = int(payload.get("quantidade") or 0)
                except (TypeError, ValueError):
                    return Response(
                        {"detail": "quantidade inválida."}, status=400
                    )
                if qtd <= 0:
                    return Response(
                        {"detail": "quantidade deve ser > 0"}, status=400
                    )

                prefixo = payload.get("descricao_prefixo", "Lote ")
                start = next_numero()
                for i in range(qtd):
                    n = start + i
                    d = f"{prefixo}{n}"
                    obj = Lote.objects.create(
                        processo=processo, numero=n, descricao=d
                    )
                    created.append(obj)

            elif isinstance(payload, dict):
                n = payload.get("numero") or next_numero()
                d = payload.get("descricao", "")
                obj = Lote.objects.create(
                    processo=processo, numero=n, descricao=d
                )
                created.append(obj)

            else:
                return Response({"detail": "Payload inválido."}, status=400)

        return Response(
            LoteSerializer(created, many=True).data,
            status=status.HTTP_201_CREATED,
        )

    @action(detail=True, methods=["patch"], url_path="lotes/organizar")
    def organizar_lotes(self, request, *args, **kwargs):
        processo = self.get_object()
        data = request.data

        with transaction.atomic():
            ordem_ids = data.get("ordem_ids")
            if isinstance(ordem_ids, list) and ordem_ids:
                qs = list(processo.lotes.filter(id__in=ordem_ids))
                id2obj = {o.id: o for o in qs}
                numero = int(data.get("inicio") or 1)
                for _id in ordem_ids:
                    obj = id2obj.get(_id)
                    if obj:
                        obj.numero = numero
                        obj.save(update_fields=["numero"])
                        numero += 1
                out = LoteSerializer(
                    processo.lotes.order_by("numero"), many=True
                ).data
                return Response(out)

            if data.get("normalizar"):
                inicio = int(data.get("inicio") or 1)
                numero = inicio
                for obj in processo.lotes.order_by("numero", "id"):
                    if obj.numero != numero:
                        obj.numero = numero
                        obj.save(update_fields=["numero"])
                    numero += 1
                out = LoteSerializer(
                    processo.lotes.order_by("numero"), many=True
                ).data
                return Response(out)

            mapa = data.get("mapa")
            if isinstance(mapa, list) and mapa:
                ids = [m.get("id") for m in mapa if m.get("id") is not None]
                qs = processo.lotes.filter(id__in=ids)
                id2obj = {o.id: o for o in qs}
                for m in mapa:
                    _id = m.get("id")
                    num = m.get("numero")
                    if _id in id2obj and isinstance(num, int) and num > 0:
                        obj = id2obj[_id]
                        obj.numero = num
                        obj.save(update_fields=["numero"])
                out = LoteSerializer(
                    processo.lotes.order_by("numero"), many=True
                ).data
                return Response(out)

        return Response({"detail": "Payload inválido."}, status=400)

    
    
# ============================================================
# 4️⃣ LOTE
# ============================================================

class LoteViewSet(viewsets.ModelViewSet):
    queryset = Lote.objects.select_related('processo').all()
    serializer_class = LoteSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_fields = ['processo']
    search_fields = ['descricao']


# ============================================================
# 5️⃣ ITEM
# ============================================================

class ItemViewSet(viewsets.ModelViewSet):
    queryset = Item.objects.select_related('processo', 'lote', 'fornecedor').all()
    serializer_class = ItemSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_fields = ['processo', 'lote', 'fornecedor']
    search_fields = ['descricao', 'unidade', 'especificacao']  # <- adiciona aqui se quiser

    def perform_create(self, serializer):
        # Garante que 'ordem' será sempre calculado no serializer
        serializer.save()

    @action(detail=True, methods=['post'], url_path='definir-fornecedor')
    def definir_fornecedor(self, request, ):
        """
        POST /api/itens/<id>/definir-fornecedor/
        Body: { "fornecedor_id": <id> }
        Vincula um fornecedor ao item.
        """
        item = self.get_object()
        fornecedor_id = request.data.get('fornecedor_id')

        if not fornecedor_id:
            return Response({'error': 'fornecedor_id é obrigatório.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            fornecedor = Fornecedor.objects.get(id=fornecedor_id)
        except Fornecedor.DoesNotExist:
            return Response({'error': 'Fornecedor não encontrado.'}, status=status.HTTP_404_NOT_FOUND)

        item.fornecedor = fornecedor
        item.save(update_fields=['fornecedor'])
        return Response({'detail': 'Fornecedor vinculado ao item com sucesso.'}, status=status.HTTP_200_OK)


# ============================================================
# 6️⃣ FORNECEDOR ↔ PROCESSO (participantes)
# ============================================================

class FornecedorProcessoViewSet(viewsets.ModelViewSet):
    queryset = FornecedorProcesso.objects.select_related('processo', 'fornecedor').all()
    serializer_class = FornecedorProcessoSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_fields = ['processo', 'fornecedor']
    # corrigido: campos existentes
    search_fields = ['fornecedor__razao_social', 'fornecedor__cnpj', 'processo__numero_processo', 'processo__numero_certame']


# ============================================================
# 7️⃣ ITEM ↔ FORNECEDOR (propostas)
# ============================================================

class ItemFornecedorViewSet(viewsets.ModelViewSet):
    queryset = ItemFornecedor.objects.select_related('item', 'fornecedor').all()
    serializer_class = ItemFornecedorSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_fields = ['item', 'fornecedor', 'vencedor']
    # corrigido: campos existentes
    search_fields = ['item__descricao', 'fornecedor__razao_social', 'fornecedor__cnpj']


# ============================================================
# 8️⃣ REORDENAÇÃO DE ITENS
# ============================================================

class ReorderItensView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, _format=None):
        """
        Body: { "item_ids": [ <id1>, <id2>, ... ] }
        Reordena itens atribuindo ordem 1..N segundo a lista enviada.
        """
        item_ids = request.data.get('item_ids', [])
        if not isinstance(item_ids, list):
            return Response({"error": "O corpo deve conter uma lista 'item_ids'."}, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            for index, item_id in enumerate(item_ids):
                try:
                    item = Item.objects.get(id=item_id)
                except Item.DoesNotExist:
                    continue
                item.ordem = index + 1
                item.save(update_fields=['ordem'])

        return Response({"status": "Itens reordenados com sucesso."}, status=status.HTTP_200_OK)


# ============================================================
# 9️⃣ USUÁRIOS E DASHBOARD
# ============================================================

class CreateUserView(generics.CreateAPIView):
    queryset = CustomUser.objects.all()
    serializer_class = UserSerializer
    permission_classes = [AllowAny]


class ManageUserView(generics.RetrieveUpdateAPIView):
    """
    GET /me/  -> retorna usuário autenticado
    PUT/PATCH -> atualiza parcialmente (JSON ou multipart para foto)
    """
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated]

    def get_object(self):
        return self.request.user

    def get_serializer_context(self):
        # necessário para montar URL absoluta da imagem
        ctx = super().get_serializer_context()
        ctx['request'] = self.request
        return ctx

    def put(self, request, *_, **__):
        serializer = self.get_serializer(self.get_object(), data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    def patch(self, request, *_, **__):
        serializer = self.get_serializer(self.get_object(), data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)


class DashboardStatsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, _request, _format=None):
        total_processos = ProcessoLicitatorio.objects.count()
        processos_em_andamento = ProcessoLicitatorio.objects.filter(situacao="Em Contratação").count()
        total_fornecedores = Fornecedor.objects.count()
        total_orgaos = Orgao.objects.count()
        total_itens = Item.objects.count()

        data = {
            'total_processos': total_processos,
            'processos_em_andamento': processos_em_andamento,
            'total_fornecedores': total_fornecedores,
            'total_orgaos': total_orgaos,
            'total_itens': total_itens,
        }
        return Response(data)


# ============================================================
# 🔟 LOGIN COM GOOGLE
# ============================================================

from google.oauth2 import id_token
from google.auth.transport import requests as google_requests
from rest_framework_simplejwt.tokens import RefreshToken
from django.conf import settings
import logging

logger = logging.getLogger(__name__)


class GoogleLoginView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        # Token vindo do front via Google Identity Services
        google_token = request.data.get("token")
        logger.info(f"Token recebido: {bool(google_token)}")

        if not google_token:
            return Response({"detail": "Token do Google ausente."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            logger.info("Validando token com Google...")
            id_info = id_token.verify_oauth2_token(
                google_token,
                google_requests.Request(),
                settings.GOOGLE_CLIENT_ID,
            )
            logger.info("Token validado com sucesso!")

            if not id_info.get("email_verified"):
                return Response({"detail": "Email não verificado pelo Google."}, status=status.HTTP_401_UNAUTHORIZED)

            email = id_info.get("email")
            nome = id_info.get("name") or ""
            picture = id_info.get("picture", "")

            if not email:
                return Response({"detail": "E-mail não fornecido pelo Google."}, status=status.HTTP_400_BAD_REQUEST)

            user, created = CustomUser.objects.get_or_create(
                email=email,
                defaults={
                    "username": email,
                    "first_name": nome.split(" ")[0],
                    "last_name": " ".join(nome.split(" ")[1:]),
                }
            )

            refresh = RefreshToken.for_user(user)
            access_token = str(refresh.access_token)

            logger.info(f"Usuário autenticado: {email}")

            return Response({
                "access": access_token,
                "refresh": str(refresh),
                "user": {
                    "id": user.id,
                    "email": user.email,
                    "name": nome,
                    "picture": picture,
                },
                "new_user": created
            }, status=status.HTTP_200_OK)

        except ValueError as e:
            logger.error(f"Token inválido: {e}")
            return Response({"detail": "Token inválido do Google."}, status=status.HTTP_401_UNAUTHORIZED)

        except Exception as e:
            logger.exception("Erro inesperado no login Google")
            return Response({"detail": "Erro no login com Google."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ContratoEmpenhoViewSet(viewsets.ModelViewSet):
    queryset = ContratoEmpenho.objects.select_related('processo').all().order_by('-criado_em', 'id')
    serializer_class = ContratoEmpenhoSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter]
    # filtros úteis para publicação/consulta
    filterset_fields = ['processo', 'ano_contrato', 'tipo_contrato_id', 'categoria_processo_id', 'receita']
    search_fields = ['numero_contrato_empenho', 'processo__numero_processo', 'ni_fornecedor']
